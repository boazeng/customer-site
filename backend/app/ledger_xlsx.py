"""בניית קובץ אקסל (.xlsx) של הכרטסת — סניף-אחר-סניף, מעוצב לצפייה והדפסה.

מקבל את אותו מבנה נתונים ש-get_customer_ledger מחזיר ומחזיר bytes של קובץ xlsx.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# צבעי המותג (תואם למסמך ה-HTML)
_BLUE = "1F3A5F"
_HEAD_BG = "EEF2F8"
_FOOT_BG = "F6F8FB"
_MONEY = '#,##0.00;[Red](#,##0.00)'

_thin = Side(style="thin", color="DDDDDD")
_border = Border(bottom=_thin)
_HEADERS = ["תאריך", "סוג תנועה", "פרטים", "חובה", "זכות", "יתרה"]


def _date(s: str) -> str:
    s = str(s or "")[:10]
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        return f"{s[8:10]}-{s[5:7]}-{s[0:4]}"
    return s


def build_ledger_xlsx(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "כרטסת"
    ws.sheet_view.rightToLeft = True   # גיליון מימין לשמאל (עברית)

    display = data.get("display_name") or data.get("custname") or ""
    branches = data.get("branches") or []

    r = 1
    ws.cell(r, 1, "כרטסת").font = Font(bold=True, size=16, color=_BLUE)
    r += 1
    ws.cell(r, 1, f"{display} · לקוח {data.get('custname','')} · "
                  f"{len(branches)} {'סניף' if len(branches) == 1 else 'סניפים'}"
            ).font = Font(size=10, color="706A60")
    r += 2

    for b in branches:
        # כותרת הסניף
        title = ("חברת " + str(b.get("company"))) if b.get("company") else "חשבון כללי"
        if b.get("branch"):
            title += " · סניף " + str(b.get("branch"))
        c = ws.cell(r, 1, title)
        c.font = Font(bold=True, size=12, color=_BLUE)
        r += 1
        sub = f"לקוח {data.get('custname','')}"
        if b.get("name"):
            sub += " — " + str(b.get("name"))
        sub += f" · {len(b.get('lines') or [])} תנועות"
        ws.cell(r, 1, sub).font = Font(size=9, color="706A60")
        r += 1

        # שורת כותרות העמודות
        for col, h in enumerate(_HEADERS, start=1):
            cell = ws.cell(r, col, h)
            cell.font = Font(bold=True, color=_BLUE)
            cell.fill = PatternFill("solid", fgColor=_HEAD_BG)
            cell.alignment = Alignment(horizontal="right")
        r += 1

        for ln in (b.get("lines") or []):
            ws.cell(r, 1, _date(ln.get("date")))
            ws.cell(r, 2, ln.get("type") or "")
            ws.cell(r, 3, ln.get("details") or ln.get("ivnum") or ln.get("fncnum") or "")
            for col, key in ((4, "debit"), (5, "credit"), (6, "balance")):
                v = ln.get(key)
                cell = ws.cell(r, col, float(v) if v not in (None, "", 0) else (0 if key == "balance" else None))
                cell.number_format = _MONEY
            for col in range(1, 7):
                ws.cell(r, col).border = _border
            r += 1

        # שורת סיכום הסניף
        ws.cell(r, 3, 'סה"כ').font = Font(bold=True, color=_BLUE)
        for col, key in ((4, "total_debit"), (5, "total_credit"), (6, "balance")):
            cell = ws.cell(r, col, float(b.get(key) or 0))
            cell.number_format = _MONEY
            cell.font = Font(bold=True, color=_BLUE)
        for col in range(1, 7):
            ws.cell(r, col).fill = PatternFill("solid", fgColor=_FOOT_BG)
        r += 2

    # יתרה כוללת
    ws.cell(r, 1, "יתרה כוללת").font = Font(bold=True, size=12, color=_BLUE)
    tot = ws.cell(r, 2, float(data.get("balance") or 0))
    tot.number_format = _MONEY
    tot.font = Font(bold=True, size=12, color=_BLUE)

    # רוחב עמודות
    for col, w in ((1, 14), (2, 22), (3, 40), (4, 14), (5, 14), (6, 16)):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
